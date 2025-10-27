from datetime import datetime
from Base.Log.log import Logger
from Base.ReadExcel.read_excel import Read_Excel
from io import BytesIO
import openpyxl


rea = Read_Excel()
log = Logger()
# 投放时间
class Placement:
    # 投放次数
    def placement_num(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        num = 0
        for i in range(2,50):
            try:
                df = ws.cell(row=i, column=1).value
                if df is None:
                    num = i - 2
                    break
            except Exception as e:
                log.error(f'读取投放次数失败,{e}')
        return num

    # 开始时间
    def beginTime(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        planum = self.placement_num(excel_path)
        beginTime_list = []
        for i in range(2,planum+2):
            try:
                df = ws.cell(row=i,column=4).value
                dt = datetime.strptime(df, "%Y %m-%d %H:%M")
                timestamp_seconds = dt.timestamp()
                timestamp_milliseconds = int(timestamp_seconds * 1000)
                log.info(f'需要投放的时间：{df}，时间戳是：{timestamp_milliseconds}')
                beginTime_list.append(timestamp_milliseconds)
            except Exception as e:
                log.error(f'读取需要投放的时间失败,{e}')
        log.info(f'开始时间:{beginTime_list}')
        return beginTime_list

    # 过期时间
    def expireTime(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        planum = self.placement_num(excel_path)
        expireTime_list = []
        for i in range(2, planum + 2):
            try:
                df = ws.cell(row=i, column=5).value
                dt = datetime.strptime(df, "%Y %m-%d %H:%M")
                timestamp_seconds = dt.timestamp()
                timestamp_milliseconds = int(timestamp_seconds * 1000)
                log.info(f'过期的时间：{df}，时间戳是：{timestamp_milliseconds}')
                expireTime_list.append(timestamp_milliseconds)
            except Exception as e:
                log.error(f'读取过期的时间失败,{e}')
        log.info(f'过期时间:{expireTime_list}')
        return expireTime_list

    # 盒子ID
    def itemId(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        planum = self.placement_num(excel_path)
        boxid_list = []
        for i in range(2, planum + 2):
            try:
                df = int(ws.cell(row=i, column=1).value)
                boxid_list.append(df)
            except Exception as e:
                log.error(f'读取盒子ID失败,{e}')
        log.info(f'盒子ID：{boxid_list}')
        return boxid_list

    # 解锁金额
    def needAmount(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        unlocktype = self.needUnlock(excel_path)
        needAmount_list = []
        for i in range(len(unlocktype)):
            try:
                if unlocktype[i] == 'true':
                    df = ws.cell(row=i+2, column=8).value
                    if df is None:
                        needAmount_list.append(20)
                    else:
                        needAmount_list.append(df)
                else:
                    needAmount_list.append(None)
            except Exception as e:
                log.error(f'读取解锁金额失败,{e}')
        log.info(f'解锁金额列表：{needAmount_list}')
        return needAmount_list

    # 是否需要解锁
    def needUnlock(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        planum = self.placement_num(excel_path)
        needUnlock_list = []
        for i in range(2, planum + 2):
            try:
                df = float(ws.cell(row=i, column=6).value)
                if df == 1:
                    needUnlock_list.append('true')
                else:
                    needUnlock_list.append('false')
            except Exception as e:
                log.error(f'读取是否需要解锁失败,{e}')
        log.info(f'是否需要解锁：{needUnlock_list}')
        return needUnlock_list

    # 是否区分时区
    def sendType(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        planum = self.placement_num(excel_path)
        sendType_list = []
        for i in range(2, planum + 2):
            try:
                df = ws.cell(row=i, column=9).value
                if df is None:
                    sendType_list.append(2)
                else:
                    sendType_list.append(df)
            except Exception as e:
                log.error(f'读取是否区分时区失败,{e}')
        log.info(f'是否区分时区：{sendType_list}')
        return sendType_list

    # 数量
    def totalCount(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        planum = self.placement_num(excel_path)
        totalCount_list = []
        for i in range(2, planum + 2):
            try:
                df = ws.cell(row=i, column=3).value
                if df is None:
                    totalCount_list.append(1)
                else:
                    totalCount_list.append(df)
            except Exception as e:
                log.error(f'读取数量失败,{e}')
        log.info(f'数量：{totalCount_list}')
        return totalCount_list

    # 解锁类型
    def unlockType(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        unlocktype = self.needUnlock(excel_path)
        unlockType_list = []
        for i in range(len(unlocktype)):
            try:
                if unlocktype[i] == 'true':
                    df = float(ws.cell(row=i + 2, column=7).value)
                    if df == 1:
                        unlockType_list.append('top_up')
                    else:
                        unlockType_list.append('service')
                else:
                    unlockType_list.append(None)
            except Exception as e:
                log.error(f'读取解锁类型失败,{e}')
        log.info(f'解锁类型：{unlockType_list}')
        return unlockType_list

    # 用户列表
    def userIdList(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        planum = self.placement_num(excel_path)
        userIdList_list = []
        for i in range(2, planum + 2):
            try:
                df = ws.cell(row=i, column=2).value.split(',')
                userIdList_list.append(df)
            except Exception as e:
                log.error(f'读取用户列表失败,{e}')
        log.info(f'用户列表：{userIdList_list}')
        return userIdList_list

    def Authorization(self,excel_path):
        virtual_file = BytesIO(excel_path)
        pla = openpyxl.load_workbook(virtual_file, data_only=True)
        ws = pla.active
        authorization = ws.cell(row=2, column=10).value
        log.info(f'用户Authorization：{authorization}')
        return authorization
